import openpyxl, pprint

workbook=openpyxl.load_workbook("censuspopdata.xlsx")
informacion=workbook["Population by Census Tract"]

ddbb={
    "nación":"United States of America",
    "estados":{}
}

for i in range(2,informacion.max_row+1):
    tract=informacion.cell(i, 1).value
    nombreEstado=informacion.cell(i, 2).value
    condado=informacion.cell(i, 3).value
    poblacion=informacion.cell(i, 4).value

    if nombreEstado not in ddbb["estados"]:
        ddbb["estados"][nombreEstado]={
            "condados":{},
            "poblacion_estado":0,
            "condados_estado":0
        }

    ddbb["estados"][nombreEstado]["poblacion_estado"]+=poblacion


    if condado not in ddbb["estados"][nombreEstado]["condados"]:
        ddbb["estados"][nombreEstado]["condados_estado"] += 1
        ddbb["estados"][nombreEstado]["condados"][condado]={
            "poblacion_condado":0,
            "tracs_condado":0
        }

    ddbb["estados"][nombreEstado]["condados"][condado]["poblacion_condado"]+=poblacion
    ddbb["estados"][nombreEstado]["condados"][condado]["tracs_condado"]+=1


pprint.pp(ddbb)