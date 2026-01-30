import openpyxl as ox

workbook= ox.load_workbook("Libro1.xlsx")
print(type(workbook))
hojas=workbook.sheetnames
print(hojas)

notas_sheet=workbook["Hoja1"]
print(notas_sheet)

print(f"Filas:{notas_sheet.min_row}")
print(f"Filas:{notas_sheet.max_row}")
print(f"Columnas:{notas_sheet.min_column}")
print(f"Columnas:{notas_sheet.max_column}")

for i in range(2,5):
    media=0
    ejercicios=0
    for j in range(2,5):
        ejercicio=notas_sheet.cell(i,j)
        ejercicios+=ejercicio.value
    ejercicios=ejercicios/3
    notas=ejercicios+notas_sheet.cell(i,5).value
    media+=notas/2
    print(f"{notas_sheet.cell(i,1).value} tiene un {media} de media")




'''hojaEnUso=workbook.active
print(hojaEnUso)

alumno=notas_sheet["A2"]
print(notas_sheet["A2"].value)


if alumno.value=="Paco":
    print("Yes")
    print(alumno.coordinate)

custom_sheet=workbook["NombreCustom"]
cel=custom_sheet["AA1"]
'''